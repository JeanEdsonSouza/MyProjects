from PySimpleGUI import PySimpleGUI as sg
import sqlite3 as sq
import openpyxl as op
import os
import pandas as pd
from datetime import date
from datetime import datetime as dt
import smtplib as sm
import mimetypes
import email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText as tx
from email.mime.application import MIMEApplication as app
#função data
sg.ChangeLookAndFeel('black')
now = date.today()
h = dt.now()
data = now.strftime('%d/%m/%y')
sg.theme('DarkBlue1')
tabel = pd.read_excel("chepe.xlsx")

def begin():
    layout =[
        [sg.Image('Entrada.png',background_color='white')],
        [sg.Button("Entrar",size=(10,1) ,key = 'entrar')]
    ]
    return sg.Window("Begin",titlebar_text_color='black', layout = layout,background_color='white',icon= 'logo.ico',finalize=True)


def Login():
    layout = [
        [sg.Text(f"bem-vindo(a)", text_color="black",background_color='white'),sg.Text("                                                                                     ",text_color='white',background_color='white'), sg.Text(f"{data}",text_color="black",background_color='white')],
        [sg.Text("Email: ",text_color="blue",background_color='white'),sg.Input(key='user1',background_color='blue',justification='left' ,size = (60,1))],
        [sg.Text("Senha: ", text_color="red",background_color='white'),sg.Input(key ='senha1',background_color='blue' ,password_char='*',size= (60,1))],
        [sg.Text("Lembre de mim", text_color='black',background_color='white'),sg.Checkbox("",text_color='black',background_color=('white'),key = 'lb')],
        [sg.Button('Login',key = 'login',size = (15,1)),sg.Text("                       ",text_color='white',background_color='white'),
         sg.Text('''        App by JEAZ 
        Cnpj:47.671.066/00001-77''',text_color="black",background_color='white')],
    ]
    return sg.Window("Login",titlebar_text_color='black',layout = layout,background_color='white',icon='logo.ico',finalize = True)

def Menu():
    layout = [
        [sg.Text("        ",background_color='white'),sg.Image('back.png',background_color='white',)],
        [sg.Button('Contagem Diaria', key = 'contagem', size = (40,3))], 
        [sg.Button("Ver_tabela", key = 'estoque',size = (40,3))],
    ]
    return sg.Window("Menu",titlebar_text_color='black',layout = layout,icon='logo.ico',finalize = True,background_color='white')

def Contagem():
    layout = [
        [sg.Text('''  Welcome to the page "Couting page"
                    here you will do your daily of chepe pallets 
                    and send it to the person responsible for administration.
                    Below you will see 2 fields:
                    1st pallet in use
                    2nd empty pallets
                    Enter numbers onlyany (letter or empty space will 
                    cause an error and the application will close.)
                    After entering the data click save and then send''',justification='center',text_color="black",background_color='white')],
        [sg.Text("Quant.Paletes Cheios: ",text_color='Black', background_color='white'), sg.Input(key = 'chepeC',background_color='DarkBlue')],
        [sg.Text("Quant.Paletes Vazios: ",text_color='Black', background_color='white'), sg.Input(key = 'chepeV',background_color='Dark Blue')],
        [sg.Button('SALVAR', key = 'salvar'), sg.Button('Enviar', key = 'enviar'),sg.Text("        ",background_color='white'), sg.Button('Voltar', key = 'voltar')],
    ]
    return sg.Window("Contagem",titlebar_text_color='black', layout = layout, icon='logo.ico',finalize = True,background_color='white')
def tabela():
    data_values = [f'{tabel}']
    data_headings = ['']
    data_values.append([f'{tabel}'])
    data_cols_width = [5, 8, 35, 35]
    tab5_layout = [
    [sg.Button("Select"), sg.Button("Salvar"),sg.Button('Select Linha', key='_selectlinha_'),sg.Button("back",key = 'back')],
    [sg.Table(values=data_values, headings=data_headings,
                            max_col_width=65,
                            col_widths=data_cols_width,
                            auto_size_columns=True,
                            justification='left',
                            num_rows=100, key='_filestable_')],
    
    ]
    
    return sg.Window("Estoque",titlebar_text_color='black',layout = tab5_layout, icon='logo.ico',finalize = True,background_color='white',size = (640,480))

entrada,login,home,contagem,storage= begin(),None,None,None,None

while True:
    window,event,values = sg.read_all_windows()
    #Alternar entre as janelas
    if window == entrada and event == 'entrar':
        entrada.hide()
        login = Login()
    
    if window == home and event == 'estoque':
        home.hide()
        storage = tabela()
    if window == home and event == 'contagem':
        home.hide()
        contagem = Contagem()
    #voltar para janela anterior
    if window == contagem and event == 'voltar':
        contagem.hide()
        home.un_hide()
    if window == storage and event == 'back':
        storage.hide()
        home.un_hide()
    #fechar janelas
    if window == login and event == sg.WINDOW_CLOSED:
        break
    if window == home and event == sg.WINDOW_CLOSED:
        break
    if window == contagem and event == sg.WINDOW_CLOSED:
        break
    if window == storage and event == sg.WINDOW_CLOSED:
        break
    #Logica das paginas
    if window == login and event == 'login':
        nome = values['user1']
        senha = values['senha1']
        b = sq.connect('user.db')
        c = b.cursor()
        c.execute(f'SELECT password FROM user WHERE user = "{nome}"')
        senha_db = c.fetchall()
        b.close()
        
        if senha == senha_db[0][0]:
            sg.PopupAutoClose('login feito com sucesso!',icon='logo.ico')
            login.hide()
            home = Menu()

    if window == contagem and event == 'salvar':
        cheio = str(values['chepeC'])
        vazio = str(values['chepeV'])
        total = cheio + vazio
        c = int(cheio)
        v = int(vazio)
        t = c+v
        ficheiro = op.load_workbook('chepe.xlsx')
        folha = ficheiro.active
        if folha.cell != "Null":
            folha.cell(column = 1, row = folha.max_row+1, value = c )
            folha.cell(column = 2, row = folha.max_row, value = v )
            folha.cell(column = 3, row = folha.max_row, value = t )
            folha.cell(column = 4, row = folha.max_row, value = data )
        ficheiro.save('chepe.xlsx')
    
    if window == contagem and event == 'enviar':
        
        msg = email.mime.multipart.MIMEMultipart()
        msg['Subject'] = 'Contagem Paletes'
        msg['From'] = 'jean.edson2015@gmail.com'
        msg["To"] = 'jeanedson.desouza@pepsico.com'
        
        body = tx('Envio Diario dos chepes')
        msg.attach(body)
        
        caminho = "chepe.xlsx"
        pdfname=f'{caminho}'
        fp=open(pdfname,'rb')
        anexo = app(fp.read(),_subtype="xlsx")
        fp.close()
        anexo.add_header('Content-Disposition','attachment',filename=pdfname)
        msg.attach(anexo)		
        email = 'jean.edson2015@gmail.com'
        senha = 'LEvi.010322'
        dest = 'jeanedson.desouza@pepsico.com'
        servidor = sm.SMTP('smtp.office365.com',587)
        servidor.starttls()
        servidor.login(f'{email}',f' {senha}' )
        servidor.sendmail(f'{email}',[f'{dest}'], msg.as_string())
        servidor.quit()