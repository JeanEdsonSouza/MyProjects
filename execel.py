#bibliotecas
import openpyxl as op
import os
import pandas as pd
from datetime import date
import smtplib as sm
import mimetypes
import email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText as tx
from email.mime.application import MIMEApplication as app
#função data
now = date.today()
data = now.strftime('%d/%m/%y')

#tela de bemvindo e inicio do codigo
print('''Bem-Vindo, selecione a opção desejada:
      [1]Contagem Diaria
      [2]NovaTabela
      [3]VisualizarTabela''')
e = int(input(": "))
#começo da logica para adcionar a contagem e salvar
while True:
    #contagem e save e email
    if e == 1:
        cheio = int(input("Quantidade de Cheios: "))
        vazio = int(input("Quantidade de Vazios: "))
        total = cheio + vazio
        ficheiro = op.load_workbook('chepe.xlsx')
        folha = ficheiro.active
        if folha.cell != "Null":
            folha.cell(column = 1, row = folha.max_row+1, value = cheio )
            folha.cell(column = 2, row = folha.max_row, value = vazio )
            folha.cell(column = 3, row = folha.max_row, value = total )
            folha.cell(column = 4, row = folha.max_row, value = data )
        ficheiro.save('chepe.xlsx')
        enviar = str(input("Deseja enviar a planilha [S/N]: ")).upper()
        if enviar == "S":
            msg = email.mime.multipart.MIMEMultipart()
            msg['Subject'] = 'Contagem Paletes'
            msg['From'] = 'Jean.edson2015@gmail.com'
            msg["To"] = 'Jeanedson.desouza@pepsico.com'
            
            body = tx('Envio Diario dos chepes')
            msg.attach(body)
            
            caminho = "chepe.xlsx"
            pdfname=f'{caminho}'
            fp=open(pdfname,'rb')
            anexo = app(fp.read(),_subtype="xlsx")
            fp.close()
            anexo.add_header('Content-Disposition','attachment',filename=pdfname)
            msg.attach(anexo)		
            email = str(input('Digite seu email: '))
            senha = str(input('Digite sua senha: '))
            dest = str(input("Digite o destinatario: "))
            servidor = sm.SMTP('smtp.office365.com',587)
            servidor.starttls()
            servidor.login(f'{email}',f' {senha}' )
            servidor.sendmail(f'{email}',[f'{dest}'], msg.as_string())
            servidor.quit()
            break
        else:
            break
    #criando nova tabela
    elif e == 2:
        import back
        c = back.criar_tabela
        break
    #Revisando e editando dados
    elif e == 3:

        tabela = pd.read_excel("chepe.xlsx")
        print(tabela)
        editar = str(input('''Deseja adcionar dados: 
                        [S/N]: ''')).upper()
        if editar == "S":
            os.system("cls")
            ficheiro = op.load_workbook('chepe.xlsx')
            folha = ficheiro.active
            tabela = pd.read_excel("chepe.xlsx")
            print(tabela)
            
            a = str(input("Selecione a Coluna desejada: "))
            b = str(input("Selecione a Linha desejada: "))

            for i in folha[f"{a}"] and folha[f"{b}"]:
                m = str(input(": "))
                folha[f'{a+b}'].value = f'{m}'
                ficheiro.save('chepe.xlsx')
                break

            ficheiro.save('testeM.xlsx')

            enviar = str(input("Deseja enviar a planilha [S/N]: ")).upper()
            if enviar == "S":

                msg = email.mime.multipart.MIMEMultipart()
                msg['Subject'] = 'Contagem Paletes'
                msg['From'] = 'Jean.edson2015@gmail.com'
                msg["To"] = 'Jeanedson.desouza@pepsico.com'
                
                body = tx('Envio Diario dos chepes')
                msg.attach(body)
                
                caminho = "chepe.xlsx"
                pdfname=f'{caminho}'
                fp=open(pdfname,'rb')
                anexo = app(fp.read(),_subtype="xlsx")
                fp.close()
                anexo.add_header('Content-Disposition','attachment',filename=pdfname)
                msg.attach(anexo)		
                email = str(input('Digite seu email: '))
                senha = str(input('Digite sua senha: '))
                dest = str(input("Digite o destinatario: "))
                servidor = sm.SMTP('smtp.office365.com',587)
                servidor.starttls()
                servidor.login(f'{email}',f' {senha}' )
                servidor.sendmail(f'{email}',[f'{dest}'], msg.as_string())
                servidor.quit()
                break
            else:
                break
        #fim do codigo    
        else:
            break